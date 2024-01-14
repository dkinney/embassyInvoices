# Description: This file contains all the styles used in the invoice generator.
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment, numbers

styles = {}

noSide = Side(style=None, color=None)
thinSide = Side(style='thin', color="000000")
thickSide = Side(style='thin', color="000000")

style = NamedStyle(name="invoiceTitle")
style.font = Font(bold=True, size=30)
style.alignment = Alignment(horizontal='right')
styles['invoiceTitle'] = style

style = NamedStyle(name="invoiceHeader")
style.alignment = Alignment(horizontal='right')
styles['invoiceHeader'] = style

style = NamedStyle(name="invoiceValue")
style.font = Font(bold=True)
styles['invoiceValue'] = style

style = NamedStyle(name="invoiceAmount")
style.number_format = numbers.BUILTIN_FORMATS[8]
style.alignment = Alignment(horizontal='left')
style.font = Font(bold=True)
styles['invoiceAmount'] = style

style = NamedStyle(name="highlighted")
style.fill = PatternFill('solid', fgColor = '00FFFF00')
styles['highlighted'] = style

style = NamedStyle(name="bordered")
style.border = Border(left=thickSide, top=thickSide, right=thickSide, bottom=thickSide)
styles['bordered'] = style

style = NamedStyle(name="focused")
style.fill = PatternFill('solid', fgColor = '00FFFF00')
style.font = Font(bold=True, size=12)
style.border = Border(left=thickSide, top=thickSide, right=thickSide, bottom=thickSide)
styles['focused'] = style

style = NamedStyle(name="defaultCell")
styles['defaultCell'] = style

style = NamedStyle(name="textCell")
style.alignment = Alignment(horizontal='left')
styles['textCell'] = style

style = NamedStyle(name="boldTextCell")
style.alignment = Alignment(horizontal='left')
style.font = Font(bold=True)
styles['boldTextCell'] = style

style = NamedStyle(name="textCellBorder")
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
style.alignment = Alignment(horizontal='left')
styles['textCell'] = style

style = NamedStyle(name="currencyCell")
style.number_format = numbers.BUILTIN_FORMATS[8]
styles['currencyCell'] = style

style = NamedStyle(name="currencyCellBorder")
style.number_format = numbers.BUILTIN_FORMATS[8]
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
styles['currencyCellBorder'] = style

style = NamedStyle(name="currencyCellTotal")
style.number_format = numbers.BUILTIN_FORMATS[8]
style.font = Font(bold=True, size=11)
styles['currencyCellTotal'] = style

style = NamedStyle(name="contractCurrency")
style.number_format = numbers.BUILTIN_FORMATS[8]
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
style.alignment = Alignment(horizontal='left')
styles['contractCurrency'] = style

style = NamedStyle(name="contractTitle")
style.alignment = Alignment(horizontal='left')
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
styles['contractTitle'] = style

style = NamedStyle(name="highlightedCurrency")
style.fill = PatternFill('solid', fgColor = '00FFFF00')
style.number_format = numbers.BUILTIN_FORMATS[8]
styles['highlightedCurrency'] = style

style = NamedStyle(name="numberCell")
style.number_format = '#,##0.0;[Red]-#,##0.0;-'
styles['numberCell'] = style

style = NamedStyle(name="numberCellBorder")
style.number_format = '#,##0.0;[Red]-#,##0.0;-'
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
styles['numberCellBorder'] = style

style = NamedStyle(name="numberCellTotal")
style.number_format = '#,##0.0;[Red]-#,##0.0;-'
style.font = Font(bold=True, size=11)
styles['numberCellTotal'] = style

style = NamedStyle(name="numberHighlighted")
style.fill = PatternFill('solid', fgColor = '00FFFF00')
style.number_format = numbers.BUILTIN_FORMATS[38]
styles['numberHighlighted'] = style

style = NamedStyle(name="numberFocused")
style.fill = PatternFill('solid', fgColor = '00FFFF00')
style.number_format = numbers.BUILTIN_FORMATS[38]
style.border = Border(left=thickSide, top=thickSide, right=thickSide, bottom=thickSide)
styles['numberFocused'] = style

style = NamedStyle(name="dateCell")
style.number_format = numbers.BUILTIN_FORMATS[14]
styles['dateCell'] = style

style = NamedStyle(name="percentageCell")
# style.number_format = numbers.BUILTIN_FORMATS[10]
style.number_format = '0%'
styles['percentageCell'] = style

style = NamedStyle(name="noBorder")
style.border = Border(left=noSide, top=noSide, right=noSide, bottom=noSide)
styles['noBorder'] = style

style = NamedStyle(name="summaryTitle")
style.alignment = Alignment(horizontal='center')
style.border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
style.font = Font(bold=True, size=11)
style.fill = PatternFill('solid', fgColor = '00000000')
style.font.color = '00FFFFFF'
styles['summaryTitle'] = style

style = NamedStyle(name="signatureLine")
style.border = Border(bottom=thinSide)
style.font = Font(bold=True, size=16)
styles['signatureLine'] = style

style = NamedStyle(name="hoursTitle")
style.fill = PatternFill('solid', fgColor = '00000000')
style.alignment = Alignment(horizontal='center')
style.font = Font(bold=True, size=18)
style.font.color = '00FFFFFF'
styles['hoursTitle'] = style

style = NamedStyle(name="hoursSubtitle")
style.fill = PatternFill('solid', fgColor = '00000000')
style.alignment = Alignment(horizontal='center')
style.font = Font(bold=True, size=14)
style.font.color = '00FFFFFF'
styles['hoursSubtitle'] = style

style = NamedStyle(name="invoiceSummaryText")
style.font = Font(bold=True, size=14)
styles['invoiceSummaryText'] = style

style = NamedStyle(name="invoiceSummaryNumber")
style.number_format = '#,##0.0;[Red]-#,##0.0;-'
style.font = Font(bold=True, size=14)
styles['invoiceSummaryNumber'] = style

style = NamedStyle(name="invoiceSummaryCurrency")
style.number_format = numbers.BUILTIN_FORMATS[8]
style.font = Font(bold=True, size=14)
styles['invoiceSummaryCurrency'] = style