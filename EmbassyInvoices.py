#!/usr/local/bin/python
import glob
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from Config import Config
from EmployeeTime import EmployeeTime
from BillingRates import BillingRates
from InvoiceStyles import styles
from InvoiceFormat import formatInvoiceTab, formatCostsTab, formatHoursTab, formatHoursDetailsTab, formatDetailTab, formatSummaryTab, formatPostDetails

config = Config()

Regions = {}

versionString = f'v{config.data["version"]}'
CountryApprovers = config.data['approvers']

# get region dictionary from config and swap keys and values
# to make it easy to look up the region name from the CLIN
for region in config.data['regions']:
	Regions[config.data['regions'][region]] = region

def getUniquifier(pattern, type=None, region=None, year=None, monthName=None):
	# look for previous instances of the status file
	patternValues = re.findall(r'(.*)-(.*)-(.*)-(.*)', pattern)
	patternType = patternValues[0][0]
	patterRegion = patternValues[0][1]
	patternYear = patternValues[0][2]
	patternMonthhName = patternValues[0][3]

	statusFiles = glob.glob(pattern + '*.xlsx')

	version = 0
	for file in statusFiles:
		for vals in re.findall(r'(.*)-(.*)-(.*)-(.*)-(\d+)', file):
			typeMatches = type is None or type is not None and vals[0] == type
			regionMatches = region is None or region is not None and vals[1] == region
			yearMatches = year is None or year is not None and vals[2] == year
			monthMatches = monthName is None or monthName is not None and vals[3] == monthName

			if typeMatches and regionMatches and yearMatches and monthMatches:
				thisVersion = int(vals[len(vals)-1])
				version = max(version, thisVersion)

	version += 1
	return version

def processActivityFromFile(filename, startInvoiceNumber=0):
	print(f'Processing activity from {filename}...')
	invoiceSummary = []

	if filename is None:
		print(f'Usage: {sys.argv[0]} <billing activity file>')
		return invoiceDetail

	activity = EmployeeTime(filename, verbose=False)
	effectiveDate = activity.dateEnd
	billingRates = BillingRates(effectiveDate=effectiveDate, verbose=False)
	activity.joinWith(billingRates)

	# do we have any activity for a clin other than 001 or 002?
	confirm = activity.data[~activity.data['CLIN'].isin(['001', '002'])]

	if not confirm.empty:
		print('\nERROR ----------------------------------------')
		print(f'Employees that did not join with billing rates:')
		print(confirm['EmployeeName'].unique())
		print('ERROR ----------------------------------------')

	# print('Time: ', activity.dateStart, ' - ', activity.dateEnd)

	startYear = activity.dateStart.strftime("%Y")
	startMonth = activity.dateStart.strftime("%m")

	locationInfo = activity.locationsByCLIN()

	# print('Location info: ', locationInfo)

	invoiceNumberValue = startInvoiceNumber

	for clin in locationInfo.keys():
		region = Regions[clin]

		##########################################################
		# Labor
		##########################################################

		reportType = 'Labor'
		pattern = f'{reportType}-{region}-{startYear}-{startMonth}'
		uniquifier = getUniquifier(pattern, type=reportType, region=region, year=startYear, monthName=startMonth)
		outputFile = f'{pattern}-{uniquifier:02d}.xlsx'
		
		sheetInfo = {}

		with pd.ExcelWriter(outputFile) as writer:
			for location in locationInfo[clin]:
				sheetName = f'Labor-{location}'
				summaryStartRow = 22
				rowsToSum = []

				data = activity.groupedForInvoicing(clin=clin, location=location)
				summary = pd.DataFrame(columns=['SubCLIN', 'Description', 'EmployeeName', 'Hours', 'Rate', 'Amount'])
				summary.loc[0] = ['', f'Totals for {location}', '', data['Hours'].sum(), '', data['Amount'].sum()]

				for subCLIN in data['SubCLIN'].unique():
					clinData = data[data['SubCLIN'] == subCLIN]
					rows = clinData.shape[0]
					clinData.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)
					rowsToSum.append((summaryStartRow + 1, summaryStartRow + rows))
					summaryStartRow = summaryStartRow + rows + 2

				summary.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)

				# invoiceNumber = laborInvoiceNumber + CountryCodes[location]
				invoiceNumber = f'SD-{invoiceNumberValue:04d}'
				invoiceNumberValue += 1
				invoiceAmount = data['Amount'].sum()

				# startMonthName = activity.dateStart.strftime('%b')
				# endMonthName = activity.dateEnd.strftime('%b')
				billingPeriod = activity.billingPeriod()

				invoiceDetail = {
					# 'description': f'{activity.dateStart.strftime("%B")} {startYear}',
					# 'region': location,
					'filename': outputFile,
					'type': 'Labor',
					'invoiceNumber': invoiceNumber,
					'taskOrder': f'Labor-{location}',
					'billingPeriod': billingPeriod,
					'invoiceAmount': invoiceAmount,
					'rowsToSum': rowsToSum
				}

				sheetInfo[sheetName] = invoiceDetail
				invoiceSummary.append(invoiceDetail)
	
		##########################################################
		# Hours Report (for signatures)
		##########################################################
				
		reportType = 'Hours'
		pattern = f'{reportType}-{region}-{startYear}-{startMonth}'
		uniquifier = getUniquifier(pattern, type=reportType, region=region, year=startYear, monthName=startMonth)
		hoursOutputFile = f'{pattern}-{uniquifier:02d}.xlsx'

		with pd.ExcelWriter(hoursOutputFile) as writer:
			for location in locationInfo[clin]:
				sheetName = f'Hours-{location}'
				# print(f'Writing hours for {location} into {sheetName}...')
				# print(data)
				data = activity.groupedForHoursReport(clin=clin, location=location)
				data.to_excel(writer, sheet_name=sheetName, startrow=0, startcol=0, header=True, index=False)

				sheetName = f'Details-{location}'
				details = activity.byDate(clin=clin, location=location)
				details.drop(columns=['State', 'Region', 'Holiday', 'Vacation', 'Bereavement', 'HoursReg', 'HoursOT', 'SubCLIN'], inplace=True)
				# rename some columns for space
				details.rename(columns={
					'EmployeeName': 'Name',
					'LocalHoliday': 'Local Hol',
					'On-callOT': 'On-call OT',
					'ScheduledOT': 'Sched OT',
					'UnscheduledOT': 'Unschd OT',
					'HoursTotal': 'Subtotal',
				}, inplace=True)
				details.to_excel(writer, sheet_name=sheetName, startrow=0, startcol=0, header=True, index=False)

		hoursWorkbook = load_workbook(hoursOutputFile)
		for styleName in styles.keys():
			hoursWorkbook.add_named_style(styles[styleName])

		for location in locationInfo[clin]:
			worksheet = hoursWorkbook[f'Hours-{location}']
			# invoiceNumber = laborInvoiceNumber + CountryCodes[location]
			
			formatHoursTab(worksheet, 
				  approvers=CountryApprovers[location], 
				  locationName=location, billingFrom=activity.billingPeriod())
			
			worksheet = hoursWorkbook[f'Details-{location}']
			formatHoursDetailsTab(worksheet, locationName=location, billingFrom=activity.billingPeriod())
		
		hoursWorkbook.save(hoursOutputFile)
		
		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])

		for key in sheetInfo.keys():
			worksheet = workbook[key]
			info = sheetInfo[key]
			formatInvoiceTab(worksheet, info)

		workbook.save(outputFile)

		##########################################################
		# PostHazard Costs
		##########################################################

		reportType = 'Post'
		pattern = f'{reportType}-{region}-{startYear}-{startMonth}'
		uniquifier = getUniquifier(pattern, type=reportType, region=region, year=startYear, monthName=startMonth)
		outputFile = f'{pattern}-{uniquifier:02d}.xlsx'

		sheetInfo = {}
		firstRow = 3
		spaceToSummary = 4

		with pd.ExcelWriter(outputFile) as writer:
			sheetName = f'PostHazard-{region}'

			costs = activity.postByCountry(clin=clin)

			post = activity.postSummaryByCity(clin=clin)
			numPostRows = post.shape[0]

			postDetails = activity.groupedForPostReport(clin=clin)
			numPostDetailRows = postDetails.shape[0]

			hazard = activity.hazardSummaryByCity(clin=clin)
			numHazardRows = hazard.shape[0]

			hazardDetails = activity.groupedForHazardReport(clin=clin)
			numHazardDetailRows = hazardDetails.shape[0]

			invoiceAmount = costs['Total'].sum()

			if invoiceAmount > 0:
				summaryStartRow = 22
				rowsToSum = []

				rows = costs.shape[0]
				costs.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)
				rowsToSum.append((summaryStartRow + 1, summaryStartRow + rows))
				summaryStartRow = summaryStartRow + rows + spaceToSummary

				invoiceNumber = f'SD-{invoiceNumberValue:04d}'
				invoiceNumberValue += 1		

				startMonthName = activity.dateStart.strftime('%b')
				endMonthName = activity.dateEnd.strftime('%b')
				billingPeriod = f'{activity.dateStart.day} {startMonthName} {activity.dateStart.year} - {activity.dateEnd.day} {endMonthName} {activity.dateEnd.year}'

				invoiceDetail = {
					# 'description': f'{activity.dateStart.strftime("%B")} {startYear}',
					# 'region': region,
					'filename': outputFile,
					'type': 'Post',
					'invoiceNumber': invoiceNumber,
					'taskOrder': f'Post-{region}',
					'billingPeriod': billingPeriod,
					'invoiceAmount': invoiceAmount,
					'rowsToSum': rowsToSum, 
					'postRows': numPostRows,
					'postDetailRows': numPostDetailRows,
					'hazardRows': numHazardRows,
					'hazardDetailRows': numHazardDetailRows
				}

				sheetInfo[sheetName] = invoiceDetail
				invoiceSummary.append(invoiceDetail)

			sheetName = f'PostHazard-{region}-Post'
			postDetails.to_excel(writer, sheet_name=sheetName, startrow=firstRow, startcol=0, header=True, index=False)
			post.to_excel(writer, sheet_name=sheetName, startrow=numPostDetailRows + firstRow + spaceToSummary, startcol=5, header=False, index=False)

			if numHazardDetailRows > 0:
				sheetName = f'PostHazard-{region}-Hazard'
				hazardDetails.to_excel(writer, sheet_name=sheetName, startrow=firstRow, startcol=0, header=True, index=False)
				hazard.to_excel(writer, sheet_name=sheetName, startrow=numHazardDetailRows + firstRow + spaceToSummary, startcol=5, header=False, index=False)

		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])
		
		for key in sheetInfo.keys():
			worksheet = workbook[key]
			info = sheetInfo[key]
			formatCostsTab(worksheet, info)

			detailSheetName = f'PostHazard-{region}-Post'
			worksheet = workbook[detailSheetName]
			postTitle = f'{region} Post {activity.dateStart.strftime("%B")} {startYear}'
			formatPostDetails(worksheet, postTitle, firstRow, info['postDetailRows'], spaceToSummary, info['postRows'])

			if info['hazardDetailRows'] > 0:
				detailSheetName = f'PostHazard-{region}-Hazard'
				worksheet = workbook[detailSheetName]
				postTitle = f'{region} Hazard {activity.dateStart.strftime("%B")} {startYear}'
				formatPostDetails(worksheet, postTitle, firstRow, info['hazardDetailRows'], spaceToSummary, info['hazardRows'])

		workbook.save(outputFile)

		##########################################################
		# Details
		##########################################################

		reportType = 'Details'
		pattern = f'{reportType}-{region}-{startYear}-{startMonth}'
		uniquifier = getUniquifier(pattern, type=reportType, region=region, year=startYear, monthName=startMonth)
		outputFile = f'{pattern}-{uniquifier:02d}.xlsx'

		# There is only one tab in the workbook
		sheetName = f'Details-{region}'

		with pd.ExcelWriter(outputFile) as writer:
			details = activity.groupedForDetailsReport(clin=clin)		
			details.to_excel(writer, sheet_name=sheetName, startrow=0, startcol=0, header=True)
			
		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])
		
		worksheet = workbook[sheetName]
		formatDetailTab(worksheet)
		workbook.save(outputFile)
	
	return invoiceSummary

def showResult(resultDictionary):
	result = pd.DataFrame(resultDictionary)
	result.drop(columns=['billingPeriod', 'rowsToSum', 'postRows', 'postDetailRows', 'hazardRows', 'hazardDetailRows'], inplace=True)
	print(result)

if __name__ == '__main__':
	import sys

	startInvoiceNumber = 123

	if len(sys.argv) < 2:
		print(f'Usage: {sys.argv[0]} <billing activity file>')
		sys.exit(1)

	processed = []

	for filename in sys.argv[1:]:
		result = processActivityFromFile(filename, startInvoiceNumber=startInvoiceNumber)
		
		showResult(result)

		for item in result:
			processed.append(item)

	invoices = pd.DataFrame(processed)

	# Drop the rowsToSum column if it exists
	invoices.drop(columns=['rowsToSum', 'postRows', 'postDetailRows', 'hazardRows', 'hazardDetailRows'], inplace=True)
	
	# if 'rowsToSum' in invoices.columns:
	# 	invoices.drop(columns=['rowsToSum'], inplace=True)

	invoices.rename(columns={
		# 'description': 'Description',
		# 'region': 'Region',
		'filename': 'Filename',
		'type': 'Type',
		'invoiceNumber': 'Invoice Number',
		'taskOrder': 'Task Order',
		'billingPeriod': 'Billing Period',
		'invoiceAmount': 'Invoice Amount'
	}, inplace=True)

	# print(invoices)

	now = pd.Timestamp.now().strftime("%Y%m%d%H%M")

	outputFile = f'Summary-{now}.xlsx'

	# if 'Billing Period' in invoices.columns:
	invoices.drop(columns=['Billing Period'], inplace=True)

	with pd.ExcelWriter(outputFile) as writer:
		invoices.to_excel(writer, sheet_name='Summary', startrow=0, startcol=0, header=True)

	# Apply formatting in place
	workbook = load_workbook(outputFile)

	for styleName in styles.keys():
		workbook.add_named_style(styles[styleName])
		
	worksheet = workbook['Summary']
	formatSummaryTab(worksheet)
	workbook.save(outputFile)
	