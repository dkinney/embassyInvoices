import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Alignment

from InvoiceStyles import styles

from Config import Config
config = Config()

contractNumber = config.data['contractNumber']
dataStyles = config.data['dataStyles']

processingDate = datetime.datetime.now().strftime('%d %b %Y')

thinSide = Side(style='thin', color="000000")
thickSide = Side(style='thin', color="000000")
yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
blue = PatternFill(start_color='21A7F2', end_color='21A7F2', fill_type='solid')
lightBlue = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

def styleColumn(worksheet, column, type, rowStart = None, rowStop = None):
    start = 0 if rowStart is None else rowStart
    stop = worksheet.max_row if rowStop is None else rowStop
    style = 'defaultCell'
    width = 12

    if dataStyles.get(type) is not None:
        style = dataStyles[type]['style']
        width = dataStyles[type]['width']
    else:
        print(f'style for {type} not found')

    # print(f'styleColumn({column}, {type}, {start}, {stop})')

    for row in range(start, stop):
        worksheet[column][row].style = style

    if rowStart is None:
        worksheet.column_dimensions[column].width = width
    else:
        worksheet[column + str(rowStart + 1)].style = 'summaryTitle'

def highlightRow(worksheet, row, color = blue, colStart = None, colStop = None):
    start = 1 if colStart is None else colStart
    stop = worksheet.max_column + 1 if colStop is None else colStop

    for col in range(start, stop):
        worksheet.cell(row=row, column=col).fill = color

def styleRow(worksheet, row, style):
    for column in range(1, worksheet.max_column):
        worksheet.cell(row=row, column=column).style = style

def columnFunction(worksheet, column, function, amountType, dataStart = None, dataStop = None, top = False):
    if amountType == 'currency':
        style = 'currencyCellTotal'
    elif amountType == 'number':
        style = 'numberCellTotal'
    else:
        style = 'defaultCell'

    start = 1 if dataStart is None else dataStart
    stop = worksheet.max_row if dataStop is None else dataStop
    totalCell = column + str(stop + 1)

    if top:
        # the start must be at least 2
        if start == 1:
            start += 1

        totalCell = column + '1'

    worksheet[totalCell] = '=SUBTOTAL(' + function +', ' + column + str(start) + ':' + column + str(stop) + ')'
    worksheet[totalCell].style = style

def sumColumn(worksheet, column, amountType, dataStart = None, dataStop = None, top = False):
    columnFunction(worksheet, column, '109', amountType, dataStart, dataStop, top)

def formatInvoiceTab(worksheet, sheetInfo):
    worksheet.delete_cols(1, 1)

    styleColumn(worksheet, 'A', 'SubCLIN')
    styleColumn(worksheet, 'B', 'Description')
    styleColumn(worksheet, 'C', 'Name')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Rate')
    styleColumn(worksheet, 'F', 'Total')

    rowsToSum = sheetInfo['rowsToSum']

    # add SUM() formulas
    for row in rowsToSum:
        sumColumn(worksheet, 'D', 'number', row[0], row[1])
        sumColumn(worksheet, 'F', 'currency', row[0], row[1])
        
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            for r in range(row[0]-1, row[1]):
                worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
    
    worksheet[f'B{worksheet.max_row}'].style = 'invoiceSummaryText'
    worksheet[f'D{worksheet.max_row}'].style = 'invoiceSummaryNumber'
    worksheet[f'F{worksheet.max_row}'].style = 'invoiceSummaryCurrency'

    logo = openpyxl.drawing.image.Image('logo-MEC.png')
    worksheet.add_image(logo, 'A1')

    worksheet['F1'] = 'Invoice'
    worksheet['F1'].style = 'invoiceTitle'

    worksheet['D3'] = 'Invoice Date:'
    worksheet['D3'].style = 'invoiceHeader'
    worksheet['E3'] = processingDate
    worksheet['E3'].style = 'invoiceValue'
    worksheet['D4'] = 'Invoice Number:'
    worksheet['D4'].style = 'invoiceHeader'
    worksheet['E4'] = sheetInfo['invoiceNumber']
    worksheet['E4'].style = 'invoiceValue'
    worksheet['D5'] = 'Invoice Amount:'
    worksheet['D5'].style = 'invoiceHeader'
    worksheet.merge_cells('E5:F5')
    worksheet['E5'] = sheetInfo['invoiceAmount']
    worksheet['E5'].style = 'invoiceAmount'
    worksheet['D6'] = 'Contract Number:'
    worksheet['D6'].style = 'invoiceHeader'
    worksheet['E6'] = contractNumber
    worksheet['E6'].style = 'invoiceValue'
    worksheet['D7'] = 'Task Order:'
    worksheet['D7'].style = 'invoiceHeader'
    worksheet['E7'] = sheetInfo['taskOrder']
    worksheet['E7'].style = 'invoiceValue'
    worksheet['D8'] = 'Billing From:'
    worksheet['D8'].style = 'invoiceHeader'
    worksheet['E8'] = sheetInfo['billingPeriod']
    worksheet['E8'].style = 'invoiceValue'
    worksheet['D9'] = 'Payment Terms:'
    worksheet['D9'].style = 'invoiceHeader'
    worksheet['E9'] = 'Net 30'
    worksheet['E9'].style = 'invoiceValue'

    row = 3
    worksheet['B' + str(row)] = config.data['address']['line1']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line2']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line3']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line4']; row += 1

    row += 4; toRow = row
    worksheet['A' + str(row)] = 'Bill To:'
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['billTo']['line1']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line2']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line3']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line4']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line5']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line6']; row += 1

    row += 1; instructionsRow = row
    worksheet['A' + str(row)] = 'ACH:'
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['ach']['bank']; row += 1
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['ach']['routing']; row += 1
    worksheet['B' + str(row)] = config.data['ach']['account']

    row = toRow
    worksheet['D' + str(row)] = 'Remit To:'
    worksheet['D' + str(row)].style = 'invoiceHeader'
    worksheet['E' + str(row)] = config.data['remitTo']['line1']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line2']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line3']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line4']; row += 1

    row = instructionsRow
    worksheet['D' + str(row)] = 'Invoice Questions:'
    worksheet['D' + str(row)].style = 'invoiceHeader'
    worksheet['E' + str(row)] = config.data['questions']['line1']; row += 1
    worksheet['E' + str(row)] = config.data['questions']['line2']; row += 1
    worksheet['E' + str(row)] = config.data['questions']['line3']; row += 1

    row += 1
    worksheet['A' + str(row)] = config.data['laborHeaders']['clin']
    worksheet['A' + str(row)].style = 'summaryTitle'
    worksheet['B' + str(row)] = config.data['laborHeaders']['description']
    worksheet['B' + str(row)].style = 'summaryTitle'
    worksheet['C' + str(row)] = config.data['laborHeaders']['type']
    worksheet['C' + str(row)].style = 'summaryTitle'
    worksheet['D' + str(row)] = config.data['laborHeaders']['quantity']
    worksheet['D' + str(row)].style = 'summaryTitle'
    worksheet['E' + str(row)] = config.data['laborHeaders']['rate']
    worksheet['E' + str(row)].style = 'summaryTitle'
    worksheet['F' + str(row)] = config.data['laborHeaders']['total']
    worksheet['F' + str(row)].style = 'summaryTitle'

def formatCostsTab(worksheet, sheetInfo):
    worksheet.delete_cols(1, 1)

    styleColumn(worksheet, 'A', 'SubCLIN')
    styleColumn(worksheet, 'B', 'Description')
    styleColumn(worksheet, 'C', 'SubCLIN')
    styleColumn(worksheet, 'D', 'Total')
    styleColumn(worksheet, 'E', 'Total')
    styleColumn(worksheet, 'F', 'Total')

    rowsToSum = sheetInfo['rowsToSum']

    # add SUM() formulas
    for row in rowsToSum:
        sumColumn(worksheet, 'D', 'currency', row[0], row[1])
        sumColumn(worksheet, 'E', 'currency', row[0], row[1])
        sumColumn(worksheet, 'F', 'currency', row[0], row[1])
        
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            for r in range(row[0]-1, row[1]):
                worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)
    
    worksheet[f'B{worksheet.max_row}'].style = 'invoiceSummaryText'
    worksheet[f'D{worksheet.max_row}'].style = 'invoiceSummaryCurrency'
    worksheet[f'E{worksheet.max_row}'].style = 'invoiceSummaryCurrency'
    worksheet[f'F{worksheet.max_row}'].style = 'invoiceSummaryCurrency'

    logo = openpyxl.drawing.image.Image('logo-MEC.png')
    worksheet.add_image(logo, 'A1')

    worksheet['F1'] = 'Invoice'
    worksheet['F1'].style = 'invoiceTitle'

    worksheet['D3'] = 'Invoice Date:'
    worksheet['D3'].style = 'invoiceHeader'
    worksheet['E3'] = datetime.datetime.now().strftime('%d %b %Y')
    worksheet['E3'].style = 'invoiceValue'
    worksheet['D4'] = 'Invoice Number:'
    worksheet['D4'].style = 'invoiceHeader'
    worksheet['E4'] = sheetInfo['invoiceNumber']
    worksheet['E4'].style = 'invoiceValue'
    worksheet['D5'] = 'Invoice Amount:'
    worksheet['D5'].style = 'invoiceHeader'
    worksheet.merge_cells('E5:F5')
    worksheet['E5'] = sheetInfo['invoiceAmount']
    worksheet['E5'].style = 'invoiceAmount'
    worksheet['D6'] = 'Contract Number:'
    worksheet['D6'].style = 'invoiceHeader'
    worksheet['E6'] = contractNumber
    worksheet['E6'].style = 'invoiceValue'
    worksheet['D7'] = 'Task Order:'
    worksheet['D7'].style = 'invoiceHeader'
    worksheet['E7'] = sheetInfo['taskOrder']
    worksheet['E7'].style = 'invoiceValue'
    worksheet['D8'] = 'Billing From:'
    worksheet['D8'].style = 'invoiceHeader'
    worksheet['E8'] = sheetInfo['billingPeriod']
    worksheet['E8'].style = 'invoiceValue'
    worksheet['D9'] = 'Payment Terms:'
    worksheet['D9'].style = 'invoiceHeader'
    worksheet['E9'] = config.data['paymentTerms']['line1']
    worksheet['E9'].style = 'invoiceValue'

    row = 3
    worksheet['B' + str(row)] = config.data['address']['line1']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line2']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line3']; row += 1
    worksheet['B' + str(row)] = config.data['address']['line4']; row += 1

    row += 4; toRow = row
    worksheet['A' + str(row)] = 'Bill To:'
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['billTo']['line1']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line2']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line3']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line4']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line5']; row += 1
    worksheet['B' + str(row)] = config.data['billTo']['line6']; row += 1

    row += 1; instructionsRow = row
    worksheet['A' + str(row)] = 'ACH:'
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['ach']['bank']; row += 1
    worksheet['A' + str(row)].style = 'invoiceHeader'
    worksheet['B' + str(row)] = config.data['ach']['routing']; row += 1
    worksheet['B' + str(row)] = config.data['ach']['account']

    row = toRow
    worksheet['D' + str(row)] = 'Remit To:'
    worksheet['D' + str(row)].style = 'invoiceHeader'
    worksheet['E' + str(row)] = config.data['remitTo']['line1']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line2']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line3']; row += 1
    worksheet['E' + str(row)] = config.data['remitTo']['line4']; row += 1

    row = instructionsRow
    worksheet['D' + str(row)] = 'Invoice Questions:'
    worksheet['D' + str(row)].style = 'invoiceHeader'
    worksheet['E' + str(row)] = config.data['questions']['line1']; row += 1
    worksheet['E' + str(row)] = config.data['questions']['line2']; row += 1
    worksheet['E' + str(row)] = config.data['questions']['line3']; row += 1

    row += 1
    worksheet['A' + str(row)] = config.data['postHeaders']['clin']
    worksheet['A' + str(row)].style = 'summaryTitle'
    worksheet['B' + str(row)] = config.data['postHeaders']['description']
    worksheet['B' + str(row)].style = 'summaryTitle'
    worksheet['C' + str(row)] = config.data['postHeaders']['type']
    worksheet['C' + str(row)].style = 'summaryTitle'
    worksheet['D' + str(row)] = config.data['postHeaders']['quantity']
    worksheet['D' + str(row)].style = 'summaryTitle'
    worksheet['E' + str(row)] = config.data['postHeaders']['rate']
    worksheet['E' + str(row)].style = 'summaryTitle'
    worksheet['F' + str(row)] = config.data['postHeaders']['total']
    worksheet['F' + str(row)].style = 'summaryTitle'

    # format the summary detail area

def formatDetailTab(worksheet):
    worksheet.delete_cols(1, 1)
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Name')
    styleColumn(worksheet, 'B', 'Date')
    styleColumn(worksheet, 'C', 'Location')
    styleColumn(worksheet, 'D', 'City')
    styleColumn(worksheet, 'E', 'SubCLIN')
    styleColumn(worksheet, 'F', 'Category')

    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')

    styleColumn(worksheet, 'M', 'Hours')
    styleColumn(worksheet, 'N', 'Hours')
    styleColumn(worksheet, 'O', 'Hours')
    styleColumn(worksheet, 'P', 'Hours')
    
    styleColumn(worksheet, 'Q', 'Rate')
    styleColumn(worksheet, 'R', 'Total')
    styleColumn(worksheet, 'S', 'Post Rate')
    styleColumn(worksheet, 'T', 'Total')
    styleColumn(worksheet, 'U', 'Hazard Rate')
    styleColumn(worksheet, 'V', 'Total')

    # create a table
    table = Table(displayName='Detail', ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)
    sumColumn(worksheet, 'N', 'number', start, stop, top=True)
    sumColumn(worksheet, 'O', 'number', start, stop, top=True)
    sumColumn(worksheet, 'P', 'number', start, stop, top=True)
    sumColumn(worksheet, 'R', 'currency', start, stop, top=True)
    sumColumn(worksheet, 'T', 'currency', start, stop, top=True)
    sumColumn(worksheet, 'V', 'currency', start, stop, top=True)

    # add a fill for columns to show that they are are summary columns
    for column in ['N', 'O', 'P']:
        for row in range(1, stop):
            worksheet[column][row].fill = lightBlue

def formatPostDetails(worksheet, title, startRow, detailRows, spaceToSummary = 2, summaryRows = 1):
    styleColumn(worksheet, 'A', 'City')
    styleColumn(worksheet, 'B', 'SubCLIN')
    styleColumn(worksheet, 'C', 'Name')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Rate')
    styleColumn(worksheet, 'F', 'Total')
    styleColumn(worksheet, 'G', 'Post Rate')
    styleColumn(worksheet, 'H', 'Total')

    logo = openpyxl.drawing.image.Image('logo-MEC.png')
    worksheet.add_image(logo, 'A1')

    worksheet.merge_cells('D2:H2')
    worksheet['D2'] = title
    worksheet['D2'].style = 'invoiceSummaryText'
    worksheet['D2'].alignment = Alignment(horizontal='center')

    # add SUM() formulas
    stop = startRow + detailRows + 1
    sumColumn(worksheet, 'H', 'currency', startRow + 1, stop)

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        worksheet[f'{column}{startRow + 1}'].style = 'summaryTitle'

        for r in range(startRow, stop):
            worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)

    start = startRow + detailRows + spaceToSummary
    stop = start + summaryRows
    sumColumn(worksheet, 'H', 'currency', start + 1, stop)

    for column in ['F', 'G', 'H']:
        worksheet[f'{column}{startRow + 1}'].style = 'summaryTitle'

        for r in range(start, stop):
            worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

def formatHoursTab(worksheet, approvers=None, locationName=None, billingFrom=None):
    aboveRows = 3
    worksheet.insert_rows(1, aboveRows)

    styleColumn(worksheet, 'A', 'City')
    styleColumn(worksheet, 'B', 'SubCLIN')
    styleColumn(worksheet, 'C', 'Name')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')

    start = 2 + aboveRows
    stop = worksheet.max_row

    # add SUM() formulas
    sumColumn(worksheet, 'D', 'number', start, stop)
    sumColumn(worksheet, 'E', 'number', start, stop)
    sumColumn(worksheet, 'F', 'number', start, stop)
    sumColumn(worksheet, 'G', 'number', start, stop)
    sumColumn(worksheet, 'H', 'number', start, stop)
    sumColumn(worksheet, 'I', 'number', start, stop)
    sumColumn(worksheet, 'J', 'number', start, stop)
    sumColumn(worksheet, 'K', 'number', start, stop)

    logo = openpyxl.drawing.image.Image('logo-MEC.png')
    worksheet.add_image(logo, 'A1')

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        worksheet[column + str(start - 1)].style = 'summaryTitle'

        for r in range(start - 2, stop):
            worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)

    signaturesRow = worksheet.max_row + 2
    worksheet.merge_cells(f'B{signaturesRow}:E{signaturesRow}')
    worksheet['B' + str(signaturesRow)].style = 'signatureLine'
    worksheet['C' + str(signaturesRow)].style = 'signatureLine'
    worksheet['D' + str(signaturesRow)].style = 'signatureLine'
    worksheet['E' + str(signaturesRow)].style = 'signatureLine'

    worksheet.merge_cells(f'G{signaturesRow}:K{signaturesRow}')
    worksheet['G' + str(signaturesRow)].style = 'signatureLine'
    worksheet['H' + str(signaturesRow)].style = 'signatureLine'
    worksheet['I' + str(signaturesRow)].style = 'signatureLine'
    worksheet['J' + str(signaturesRow)].style = 'signatureLine'
    worksheet['K' + str(signaturesRow)].style = 'signatureLine'

    signaturesRow += 1
    worksheet.merge_cells(f'B{signaturesRow}:E{signaturesRow}')
    worksheet['B' + str(signaturesRow)].style = 'boldTextCell'
    worksheet['B' + str(signaturesRow)].value = approvers['MES']
    worksheet.merge_cells(f'G{signaturesRow}:K{signaturesRow}')
    worksheet['G' + str(signaturesRow)].value = approvers['COR']
    worksheet['G' + str(signaturesRow)].style = 'boldTextCell'

    worksheet['G1'] = 'Invoice Date:'
    worksheet['G1'].style = 'invoiceHeader'
    worksheet['H1'] = processingDate
    worksheet['H1'].style = 'invoiceValue'

    if locationName is not None:
        worksheet['G2'] = 'Location:'
        worksheet['G2'].style = 'invoiceHeader'
        worksheet['H2'] = locationName
        worksheet['H2'].style = 'invoiceValue'

    if billingFrom is not None:
        worksheet['G3'] = 'Billing From:'
        worksheet['G3'].style = 'invoiceHeader'
        worksheet['H3'] = billingFrom
        worksheet['H3'].style = 'invoiceValue'

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

def formatHoursDetailsTab(worksheet, locationName=None, invoiceNumber=None, billingFrom=None):
    aboveRows = 3
    worksheet.insert_rows(1, aboveRows)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'Hours')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')

    start = 2 + aboveRows
    stop = worksheet.max_row

    # add SUM() formulas
    sumColumn(worksheet, 'C', 'number', start, stop)
    sumColumn(worksheet, 'D', 'number', start, stop)
    sumColumn(worksheet, 'E', 'number', start, stop)
    sumColumn(worksheet, 'F', 'number', start, stop)
    sumColumn(worksheet, 'G', 'number', start, stop)
    sumColumn(worksheet, 'H', 'number', start, stop)
    sumColumn(worksheet, 'I', 'number', start, stop)
    sumColumn(worksheet, 'J', 'number', start, stop)

    logo = openpyxl.drawing.image.Image('logo-MEC.png')
    worksheet.add_image(logo, 'A1')

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        worksheet[column + str(start - 1)].style = 'summaryTitle'

        for r in range(start - 2, stop):
            worksheet[column + str(r + 1)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)

    worksheet['G1'] = 'Invoice Date:'
    worksheet['H1'] = processingDate
    worksheet['G1'].style = 'invoiceHeader'
    worksheet['H1'].style = 'invoiceValue'

    if locationName is not None:
        worksheet['G2'] = 'Location:'
        worksheet['G2'].style = 'invoiceHeader'
        worksheet['H2'] = locationName
        worksheet['H2'].style = 'invoiceValue'

    if billingFrom is not None:
        worksheet['G3'] = 'Billing From:'
        worksheet['G3'].style = 'invoiceHeader'
        worksheet['H3'] = billingFrom
        worksheet['H3'].style = 'invoiceValue'
    
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

def formatFullDetailsTab(worksheet):
    worksheet.delete_cols(1, 1)
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'CLIN')
    styleColumn(worksheet, 'B', 'Location')
    styleColumn(worksheet, 'C', 'City')
    styleColumn(worksheet, 'D', 'SubCLIN')
    styleColumn(worksheet, 'E', 'Category')
    styleColumn(worksheet, 'F', 'Name')

    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')

    styleColumn(worksheet, 'M', 'Hours')
    styleColumn(worksheet, 'N', 'Hours')
    styleColumn(worksheet, 'O', 'Hours')
    styleColumn(worksheet, 'P', 'Hours')

    styleColumn(worksheet, 'Q', 'Hours')
    styleColumn(worksheet, 'R', 'Hours')
    styleColumn(worksheet, 'S', 'Hours')
    
    styleColumn(worksheet, 'T', 'Rate')
    styleColumn(worksheet, 'U', 'Total')
    styleColumn(worksheet, 'V', 'Total')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)
    sumColumn(worksheet, 'N', 'number', start, stop, top=True)
    sumColumn(worksheet, 'O', 'number', start, stop, top=True)
    sumColumn(worksheet, 'P', 'number', start, stop, top=True)
    sumColumn(worksheet, 'Q', 'number', start, stop, top=True)
    sumColumn(worksheet, 'R', 'number', start, stop, top=True)
    sumColumn(worksheet, 'S', 'number', start, stop, top=True)

    sumColumn(worksheet, 'U', 'currency', start, stop, top=True)
    sumColumn(worksheet, 'V', 'currency', start, stop, top=True)

def formatSummaryTab(worksheet):
    worksheet.delete_cols(1, 1)

    styleColumn(worksheet, 'A', 'Filename')
    styleColumn(worksheet, 'B', 'Type')
    styleColumn(worksheet, 'C', 'InvoiceNumber')
    styleColumn(worksheet, 'D', 'TaskOrder')
    styleColumn(worksheet, 'E', 'Total')

    for column in ['A', 'B', 'C', 'D', 'E']:
        worksheet[column + '1'].style = 'summaryTitle'

        for r in range(1, worksheet.max_row + 1):
                worksheet[column + str(r)].border = Border(left=thinSide, top=thinSide, right=thinSide, bottom=thinSide)

    # add SUM() formulas
    start = 2
    stop = worksheet.max_row
    sumColumn(worksheet, 'E', 'currency', start, stop)

def formatDaysTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'Task Name')
    styleColumn(worksheet, 'D', 'Regular')
    styleColumn(worksheet, 'E', 'Regular')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 2
    stop = worksheet.max_row
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)

    worksheet['B1'].value = '=D1-E1'
    worksheet['B1'].style = 'numberCellTotal'

    for row in range(3, worksheet.max_row):
        if worksheet[f'D{row}'].value != worksheet[f'E{row}'].value:
            worksheet[f'D{row}'].fill = yellow
            worksheet[f'E{row}'].fill = yellow

def formatEmployeesTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Name')
    styleColumn(worksheet, 'B', 'Regular')
    styleColumn(worksheet, 'C', 'Regular')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 2
    stop = worksheet.max_row
    sumColumn(worksheet, 'B', 'number', start, stop, top=True)
    sumColumn(worksheet, 'C', 'number', start, stop, top=True)

    worksheet['A1'].value = '=B1-C1'
    worksheet['A1'].style = 'numberCellTotal'

    for row in range(3, worksheet.max_row):
        if worksheet[f'B{row}'].value != worksheet[f'C{row}'].value:
            worksheet[f'B{row}'].fill = yellow 
            worksheet[f'C{row}'].fill = yellow 
    
def formatTasksTab(worksheet):
    worksheet.insert_rows(1, 1)
    

    maxColumns = worksheet.max_column
    # print(f'maxColumns: {maxColumns}, {get_column_letter(maxColumns)}')

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'Task Name')

    for column in range(4, maxColumns + 1):
        # print(f'styling column: {get_column_letter(column)}')
        styleColumn(worksheet, get_column_letter(column), 'Regular')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(maxColumns) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 2
    stop = worksheet.max_row
    for column in range(3, maxColumns + 1):
        # print(f'summing column: {get_column_letter(column)}')
        sumColumn(worksheet, get_column_letter(column), 'number', start, stop, top=True)

    for column in range(3, 10):
        # compare this column to the corresponding TCP column
        thisColumn = get_column_letter(column)
        tcpColumn = get_column_letter(column + 7)

        # print(f'comparing {get_column_letter(column)} to {get_column_letter(tcpColumn)}')

        for row in range(3, worksheet.max_row):
            if worksheet[f'{thisColumn}{row}'].value != worksheet[f'{tcpColumn}{row}'].value:
                worksheet[f'A{row}'].fill = yellow
                worksheet[f'B{row}'].fill = yellow 
                worksheet[f'{thisColumn}{row}'].fill = yellow
                worksheet[f'{tcpColumn}{row}'].fill = yellow

    worksheet['B1'].value = '=Q1-R1'
    worksheet['B1'].style = 'numberCellTotal'

def formatActivityDataTab(worksheet):
    styleColumn(worksheet, 'A', 'Date')
    # styleColumn(worksheet, 'B', 'CLIN')
    # styleColumn(worksheet, 'C', 'Location')
    # styleColumn(worksheet, 'D', 'City')
    # styleColumn(worksheet, 'E', 'SubCLIN')
    # styleColumn(worksheet, 'F', 'Category')
    # styleColumn(worksheet, 'G', 'Description')
    # styleColumn(worksheet, 'H', 'Name')
    # styleColumn(worksheet, 'I', 'Task ID')
    # styleColumn(worksheet, 'J', 'Task Name')
    # styleColumn(worksheet, 'K', 'Hours')
    # styleColumn(worksheet, 'L', 'Rate')
    # styleColumn(worksheet, 'M', 'Rate')
    # styleColumn(worksheet, 'N', 'Rate')
    # styleColumn(worksheet, 'O', 'Rate')

    # create a table
    # table = Table(displayName=worksheet.title, ref=f'A1:O{worksheet.max_row}')
    # worksheet.add_table(table)
    # worksheet.freeze_panes = worksheet['A2']

def formatDebugTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Name')
    styleColumn(worksheet, 'B', 'Hours')
    styleColumn(worksheet, 'C', 'Hours')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')
    styleColumn(worksheet, 'M', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'B', 'number', start, stop, top=True)
    sumColumn(worksheet, 'C', 'number', start, stop, top=True)
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)
    sumColumn(worksheet, 'F', 'number', start, stop, top=True)
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)

def formatPivotTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'Hours')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:L" + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'C', 'number', start, stop, top=True)
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)
    sumColumn(worksheet, 'F', 'number', start, stop, top=True)
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)

def formatJoinedPivotTab(worksheet, taskOffset):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')

    # Intacct
    styleColumn(worksheet, 'C', 'Hours')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')
    styleColumn(worksheet, 'M', 'Hours')

    # TCP
    styleColumn(worksheet, 'N', 'Hours')
    styleColumn(worksheet, 'O', 'Hours')
    styleColumn(worksheet, 'P', 'Hours')
    styleColumn(worksheet, 'Q', 'Hours')
    styleColumn(worksheet, 'R', 'Hours')
    styleColumn(worksheet, 'S', 'Hours')
    styleColumn(worksheet, 'T', 'Hours')
    styleColumn(worksheet, 'U', 'Hours')
    styleColumn(worksheet, 'V', 'Hours')
    styleColumn(worksheet, 'W', 'Hours')
    styleColumn(worksheet, 'X', 'Hours')

    # create a table
    # ref = f'$A$2:$X${worksheet.max_row}'
    # print(f'{worksheet.title} ref: {ref}')
    # table = Table(displayName=worksheet.title, ref=ref)
    # worksheet.add_table(table)
    # worksheet.freeze_panes = worksheet['A3']

    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'C', 'number', start, stop, top=True)
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)
    sumColumn(worksheet, 'F', 'number', start, stop, top=True)
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)

    sumColumn(worksheet, 'N', 'number', start, stop, top=True)
    sumColumn(worksheet, 'O', 'number', start, stop, top=True)
    sumColumn(worksheet, 'P', 'number', start, stop, top=True)
    sumColumn(worksheet, 'Q', 'number', start, stop, top=True)
    sumColumn(worksheet, 'R', 'number', start, stop, top=True)
    sumColumn(worksheet, 'S', 'number', start, stop, top=True)
    sumColumn(worksheet, 'T', 'number', start, stop, top=True)
    sumColumn(worksheet, 'U', 'number', start, stop, top=True)
    sumColumn(worksheet, 'V', 'number', start, stop, top=True)
    sumColumn(worksheet, 'W', 'number', start, stop, top=True)
    sumColumn(worksheet, 'X', 'number', start, stop, top=True)

    for column in range(3, 14):
        # compare this column to the corresponding TCP column
        thisColumn = get_column_letter(column)
        tcpColumn = get_column_letter(column + taskOffset)

        for row in range(start, stop + 1):
            color = yellow

            # if the subtotals match, then it is merely a difference in Task
            if worksheet[f'M{row}'].value == worksheet[f'X{row}'].value:
                color = orange

            if worksheet[f'{thisColumn}{row}'].value != worksheet[f'{tcpColumn}{row}'].value:
                worksheet[f'A{row}'].fill = color
                worksheet[f'B{row}'].fill = color 
                worksheet[f'{thisColumn}{row}'].fill = color
                worksheet[f'{tcpColumn}{row}'].fill = color

    # worksheet['A1'].value = '=X1-M1'
    # worksheet['A1'].style = 'numberCellTotal'

def formatDiffsTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'Task Name')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:E" + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)

    worksheet['A1'].value = '=D1-E1'
    worksheet['A1'].style = 'numberCellTotal'

    for row in range(3, worksheet.max_row + 1):
        if worksheet[f'D{row}'].value != worksheet[f'E{row}'].value:
            worksheet[f'D{row}'].fill = yellow
            worksheet[f'E{row}'].fill = yellow

def highlightDiffs(worksheet1, worksheet2) -> bool:
    rows1 = worksheet1.max_row
    rows2 = worksheet2.max_row
    cols1 = worksheet1.max_column
    cols2 = worksheet2.max_column

    rows = max(rows1, rows2)

    if cols1 != cols2:
        print('\n\n\nThe two tabs do not have the same number of columns!')
        print(f'cols1: {cols1}, cols2: {cols2}')
        return False
    
    for row in range(1, rows):
        for col in range(1, cols1 + 1):
            if worksheet1.cell(row=row, column=col).value != worksheet2.cell(row=row, column=col).value:
                worksheet1.cell(row=row, column=col).fill = yellow
                worksheet2.cell(row=row, column=col).fill = yellow
    
    return True

def formatJoinTab(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Hours')
    styleColumn(worksheet, 'C', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'B', 'number', start, stop, top=True)
    sumColumn(worksheet, 'C', 'number', start, stop, top=True)

    worksheet['A1'].value = '=B1-C1'
    worksheet['A1'].style = 'numberCellTotal'

    for row in range(3, worksheet.max_row + 1):
        if worksheet[f'B{row}'].value != worksheet[f'C{row}'].value:
            worksheet[f'B{row}'].fill = yellow
            worksheet[f'C{row}'].fill = yellow

def formatTimeByDate(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Date')
    styleColumn(worksheet, 'B', 'Name')
    styleColumn(worksheet, 'C', 'SubCLIN')
    styleColumn(worksheet, 'D', 'State')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')
    styleColumn(worksheet, 'M', 'Hours')
    styleColumn(worksheet, 'N', 'Hours')
    styleColumn(worksheet, 'O', 'Hours')
    styleColumn(worksheet, 'P', 'Hours')
    styleColumn(worksheet, 'Q', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:P" + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)
    sumColumn(worksheet, 'F', 'number', start, stop, top=True)
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)
    sumColumn(worksheet, 'N', 'number', start, stop, top=True)
    sumColumn(worksheet, 'O', 'number', start, stop, top=True)
    sumColumn(worksheet, 'P', 'number', start, stop, top=True)
    sumColumn(worksheet, 'Q', 'number', start, stop, top=True)

    for row in range(3, stop + 1):
        if worksheet[f'D{row}'].value == "Approved":
            pass
        elif worksheet[f'D{row}'].value == "Submitted":
            highlightRow(worksheet, row, color=yellow)
        elif worksheet[f'D{row}'].value == "Draft":
            highlightRow(worksheet, row, color=gray) 
        elif worksheet[f'D{row}'].value == "Declined":
            highlightRow(worksheet, row, color=orange)  
        else:
            highlightRow(worksheet, row)

    # add a fill for columns that are non-billable
    for column in ['H', 'I', 'J']:
        for row in range(1, stop):
            worksheet[column][row].fill = gray

    # add a fill for columns to show that they are are summary columns
    for column in ['O', 'P', 'Q']:
        for row in range(1, stop):
            worksheet[column][row].fill = lightBlue

def formatTimeByEmployee(worksheet):
    worksheet.insert_rows(1, 1)

    styleColumn(worksheet, 'A', 'Name')
    styleColumn(worksheet, 'B', 'SubCLIN')
    styleColumn(worksheet, 'C', 'State')
    styleColumn(worksheet, 'D', 'Hours')
    styleColumn(worksheet, 'E', 'Hours')
    styleColumn(worksheet, 'F', 'Hours')
    styleColumn(worksheet, 'G', 'Hours')
    styleColumn(worksheet, 'H', 'Hours')
    styleColumn(worksheet, 'I', 'Hours')
    styleColumn(worksheet, 'J', 'Hours')
    styleColumn(worksheet, 'K', 'Hours')
    styleColumn(worksheet, 'L', 'Hours')
    styleColumn(worksheet, 'M', 'Hours')
    styleColumn(worksheet, 'N', 'Hours')
    styleColumn(worksheet, 'O', 'Hours')
    styleColumn(worksheet, 'P', 'Hours')

    # create a table
    table = Table(displayName=worksheet.title, ref="A2:O" + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A3']

    # add SUM() formulas
    start = 3
    stop = worksheet.max_row
    sumColumn(worksheet, 'D', 'number', start, stop, top=True)
    sumColumn(worksheet, 'E', 'number', start, stop, top=True)
    sumColumn(worksheet, 'F', 'number', start, stop, top=True)
    sumColumn(worksheet, 'G', 'number', start, stop, top=True)
    sumColumn(worksheet, 'H', 'number', start, stop, top=True)
    sumColumn(worksheet, 'I', 'number', start, stop, top=True)
    sumColumn(worksheet, 'J', 'number', start, stop, top=True)
    sumColumn(worksheet, 'K', 'number', start, stop, top=True)
    sumColumn(worksheet, 'L', 'number', start, stop, top=True)
    sumColumn(worksheet, 'M', 'number', start, stop, top=True)
    sumColumn(worksheet, 'N', 'number', start, stop, top=True)
    sumColumn(worksheet, 'O', 'number', start, stop, top=True)
    sumColumn(worksheet, 'P', 'number', start, stop, top=True)

    for row in range(3, stop + 1):
        if worksheet[f'C{row}'].value == "Approved":
            pass
        elif worksheet[f'C{row}'].value == "Submitted":
            highlightRow(worksheet, row, color=yellow)
        elif worksheet[f'C{row}'].value == "Draft":
            highlightRow(worksheet, row, color=gray)
        elif worksheet[f'C{row}'].value == "Declined":
            highlightRow(worksheet, row, color=orange)  
        else:
            highlightRow(worksheet, row)

    # add a fill for columns that are non-billable
    for column in ['G', 'H', 'I']:
        for row in range(1, stop):
            worksheet[column][row].fill = gray
    # add a fill for columns to show that they are are summary columns
    for column in ['N', 'O', 'P']:
        for row in range(1, stop):
            worksheet[column][row].fill = lightBlue

def formatEmployeeInfo(worksheet):
    styleColumn(worksheet, 'A', 'Name')
    styleColumn(worksheet, 'B', 'Number')
    styleColumn(worksheet, 'C', 'Date')
    styleColumn(worksheet, 'D', 'Title')
    styleColumn(worksheet, 'E', 'Rate')
    styleColumn(worksheet, 'F', 'Rate')
    styleColumn(worksheet, 'G', 'Location')
    styleColumn(worksheet, 'H', 'City')
    styleColumn(worksheet, 'I', 'Post Rate')
    styleColumn(worksheet, 'J', 'Hazard Rate')
    styleColumn(worksheet, 'K', 'CLIN')
    styleColumn(worksheet, 'L', 'SubCLIN')
    styleColumn(worksheet, 'M', 'Category')
    styleColumn(worksheet, 'N', 'Rate')
    styleColumn(worksheet, 'O', 'Rate')

    # create a table
    table = Table(displayName=worksheet.title, ref="A1:O" + str(worksheet.max_row))
    worksheet.add_table(table)
    worksheet.freeze_panes = worksheet['A2']

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        worksheet[f'{column}1'].style = 'summaryTitle'