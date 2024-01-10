#!/usr/local/bin/python
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from BillingActivity import BillingActivity
from InvoiceStyles import styles
from InvoiceFormat import formatInvoiceTab, formatCostsTab, formatDetailTab, formatSummaryTab

Regions = {
	'001': 'Asia',
	'002': 'Europe'
}

CountryCodes = {
	'China': 'CH',
	'Hong Kong': 'HK',
	'Vietnam': 'VN',
	'Belgium': 'BE',
	'Moldova': 'MD',
	'Russia': 'RU',
	'Ukraine': 'UA',
	'NATO': 'NATO'
}

def processActivityFromFile(filename):
	print(f'Processing activity from {filename}...')

	invoiceSummary = []

	if filename is None:
		print(f'Usage: {sys.argv[0]} <billing activity file>')
		return invoiceDetail

	activity = BillingActivity(filename, verbose=False)

	startYear = activity.dateStart.strftime("%Y")
	startMonth = activity.dateStart.strftime("%m")

	locationInfo = activity.locationsByCLIN()

	for clin in locationInfo.keys():
		region = Regions[clin]

		##########################################################
		# Labor
		##########################################################

		outputFile = f'Labor-{startYear}{startMonth}-{region}-v3.xlsx'
		laborInvoiceNumber = f'SDEL-{startYear}{startMonth}'

		sheetInfo = {}

		with pd.ExcelWriter(outputFile) as writer:
			for location in locationInfo[clin]:
				sheetName = f'Labor-{location}'
				summaryStartRow = 22
				rowsToSum = []

				data = activity.groupedForInvoicing(clin=clin, location=location)
				summary = pd.DataFrame(columns=['SubCLIN', 'Description', 'EmployeeName', 'Hours', 'Rate', 'Amount'])
				summary.loc[0] = ['', f'Totals for {location}', '', data['Hours'].sum(), '', data['Amount'].sum()]

				for subCLIN in data['SubCLIN'].unique():
					clinData = data[data['SubCLIN'] == subCLIN]
					rows = clinData.shape[0]
					clinData.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)
					rowsToSum.append((summaryStartRow + 1, summaryStartRow + rows))
					summaryStartRow = summaryStartRow + rows + 2

				summary.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)

				invoiceNumber = laborInvoiceNumber + CountryCodes[location]
				invoiceAmount = data['Amount'].sum()

				startMonthName = activity.dateStart.strftime('%b')
				endMonthName = activity.dateEnd.strftime('%b')
				billingPeriod = f'{activity.dateStart.day} {startMonthName} {activity.dateStart.year} - {activity.dateEnd.day} {endMonthName} {activity.dateEnd.year}'

				invoiceDetail = {
					'description': f'{activity.dateStart.strftime("%B")} {startYear}',
					'region': region,
					'filename': outputFile,
					'type': 'Labor',
					'invoiceNumber': invoiceNumber,
					'taskOrder': location,
					'billingPeriod': billingPeriod,
					'invoiceAmount': invoiceAmount,
					'rowsToSum': rowsToSum
				}

				sheetInfo[sheetName] = invoiceDetail
				invoiceSummary.append(invoiceDetail)
	
		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])

		for key in sheetInfo.keys():
			worksheet = workbook[key]
			info = sheetInfo[key]
			formatInvoiceTab(worksheet, info)

		workbook.save(outputFile)

		##########################################################
		# Costs
		##########################################################

		outputFile = f'Costs-{startYear}{startMonth}-{region}-v3.xlsx'
		costInvoiceNumber = f'SDEC-{startYear}{startMonth}'

		sheetInfo = {}

		with pd.ExcelWriter(outputFile) as writer:
			sheetName = f'Costs-{region}'

			costs = activity.groupedForCosts(clin=clin)
			invoiceAmount = costs['Total'].sum()

			if invoiceAmount > 0:
				summaryStartRow = 22
				rowsToSum = []

				rows = costs.shape[0]
				costs.to_excel(writer, sheet_name=sheetName, startrow=summaryStartRow, startcol=0, header=False)
				rowsToSum.append((summaryStartRow + 1, summaryStartRow + rows))
				summaryStartRow = summaryStartRow + rows + 2

				# use the first character of the region to uniquely identify this invoice
				uniqueRegion = region[0].upper()
				invoiceNumber = costInvoiceNumber + uniqueRegion

				startMonthName = activity.dateStart.strftime('%b')
				endMonthName = activity.dateEnd.strftime('%b')
				billingPeriod = f'{activity.dateStart.day} {startMonthName} {activity.dateStart.year} - {activity.dateEnd.day} {endMonthName} {activity.dateEnd.year}'

				invoiceDetail = {
					'description': f'{activity.dateStart.strftime("%B")} {startYear}',
					'region': region,
					'filename': outputFile,
					'type': 'Costs',
					'invoiceNumber': invoiceNumber,
					'taskOrder': f'ODC-{region}',
					'billingPeriod': billingPeriod,
					'invoiceAmount': invoiceAmount,
					'rowsToSum': rowsToSum
				}

				sheetInfo[sheetName] = invoiceDetail
				invoiceSummary.append(invoiceDetail)

		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])
		
		for key in sheetInfo.keys():
			worksheet = workbook[key]
			info = sheetInfo[key]
			formatCostsTab(worksheet, info)

		workbook.save(outputFile)

		##########################################################
		# Details
		##########################################################

		outputFile = f'Details-{startYear}{startMonth}-{region}-v3.xlsx'

		# There is only one tab in the workbook
		sheetName = f'Details-{region}'

		with pd.ExcelWriter(outputFile) as writer:
			details = activity.details(clin=clin)
			details.drop(columns=['CLIN'], inplace=True)
			details.to_excel(writer, sheet_name=sheetName, startrow=0, startcol=0, header=True)
			
		# Apply formatting in place
		workbook = load_workbook(outputFile)

		for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])
		
		worksheet = workbook[sheetName]
		formatDetailTab(worksheet)
		workbook.save(outputFile)
	
	return invoiceSummary

def showResult(resultDictionary):
	result = pd.DataFrame(resultDictionary)
	result.drop(columns=['rowsToSum'], inplace=True)
	print(result)

if __name__ == '__main__':
	import sys

	processed = []

	for filename in sys.argv[1:]:
		result = processActivityFromFile(filename)
		showResult(result)

		for item in result:
			processed.append(item)

	invoices = pd.DataFrame(processed)

	# Drop the rowsToSum column if it exists
	if 'rowsToSum' in invoices.columns:
		invoices.drop(columns=['rowsToSum'], inplace=True)

	invoices.rename(columns={
		'description': 'Description',
		'region': 'Region',
		'filename': 'Filename',
		'type': 'Type',
		'invoiceNumber': 'Invoice Number',
		'taskOrder': 'Task Order',
		'billingPeriod': 'Billing Period',
		'invoiceAmount': 'Invoice Amount'
	}, inplace=True)

	print(invoices)

	now = pd.Timestamp.now().strftime("%Y%m%d%H%M")

	outputFile = f'Summary-{now}.xlsx'

	with pd.ExcelWriter(outputFile) as writer:
		invoices.to_excel(writer, sheet_name='Summary', startrow=0, startcol=0, header=True)

	# Apply formatting in place
	workbook = load_workbook(outputFile)

	for styleName in styles.keys():
			workbook.add_named_style(styles[styleName])
		
	worksheet = workbook['Summary']
	formatSummaryTab(worksheet)
	workbook.save(outputFile)
	